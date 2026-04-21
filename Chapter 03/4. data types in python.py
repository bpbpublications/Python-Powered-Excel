# Data Types in Python

# Integers (int) ---------------------------------

x = 5; x
type(x)

quantity = 10; type(quantity)  #int
score = -5; type(score)  #int

# Working with large numbers
big_number = 123456789123456789123456789123456789
print(big_number)  # Works seamlessly

# Floats -----------------------------------------

x = 10/5; x
type(x)

price = 999.99; type(price) # float

float_lim = 2.123456789123456789
float_lim # upto 15 digits

0.1 + 0.2 #0.30000000000000004

# OPTIONAL (Using decimal module for precision) ------------------

from decimal import Decimal
Decimal('0.1') + Decimal('0.2')

Decimal(float_lim)
Decimal('2.123456789123456789')

## Addition
price = Decimal('10.99') 
print(price)  

## Multiplication
tax_rate = Decimal('0.0825')
price_with_tax = price * (1 + tax_rate)
print(price_with_tax)  

10.99*(1+0.08)

# Booleans (bool) ---------------------------------

in_stock = True; type(in_stock) # True
10 > 20 # False

# Strings ---------------------------------

product_name = "Laptop"; product_name
type(product_name)

score = 'Twenty'; score
type(score)

# Now, print this again!
product_name
print(product_name)

# Did you notice any difference?
# When you just type a variable like product_name, Python shows 
# a developer-friendly version (with quotes). 
# But print(product_name) displays it cleanly as output, 
# just like Excel would show cell contents to a user.

# Another example_Notice the way current directory is displayed:
import os

os.getcwd()
print(os.getcwd())    

# Typing os.getcwd() shows the path with quotes (as a string), 
# while print(os.getcwd()) displays it as plain text without quotes—same on Windows and other OS.

# Multi-line Strings
msg = '''Hi, I am Dr Nisha. I welcome you all to learn Python for Excel.
Learning Python is easy and using it with Excel is pure-fun!
Keep learning!'''
msg
print(msg)    

# Dates ---------------------------------
    
from datetime import date
order_date = date(2025, 1, 1) # A date value similar to Excel
print(order_date) # 2025-01-01

print('current_date: ', date.today()) # print today's date in yyyy-mm-dd format

# Additional Stuff -----------------------------------------

# infinity:
x = 1e308; x # 1e+308
y = x * 10; y # inf

# when numbers go beyond its limit, Python automatically switches to inf 
# when floating-point numbers overflow

x = 'inf'; print(x) # inf
type(x) # str
type(float(x)) # float

print(float('inf'))  # inf
print(1000 * float('inf'))  # Still inf
print(1 / float('inf'))  # 0.0

# --------------------------------------------------------------
# For now, you can skip these examples
# You will understand all of this in due course

# None:
x = None; x    # Nothing to show 
print(x)       # None
type(x)        # NoneType

# Importing Excel file containing blank cells, #N/A or empty string using different libraries

import pandas as pd
df = pd.read_excel('data_blank_cells.xlsx')
df
    
from openpyxl import load_workbook
wb = load_workbook('data_blank_cells.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2, values_only=True):
    # Each 'row' is a tuple of cell values
    print(row)

# Nan:
import math

x = math.nan; x  # nan
print(x)         # nan
type(x)          # float

import numpy as np
x = np.nan; x    # nan
print(x)         # nan
type(x)          # float