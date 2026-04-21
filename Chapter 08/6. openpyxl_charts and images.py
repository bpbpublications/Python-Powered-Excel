from openpyxl import load_workbook
file = "product_data.xlsx"
wb = load_workbook(file)
ws = wb.active


# Bar Chart

from openpyxl.chart import BarChart, Reference


# Create Bar Chart for Units Sold
chart = BarChart()
chart.title = "Units Sold by Product"
chart.x_axis.title = "Product"
chart.y_axis.title = "Units Sold"

categories = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=ws.max_row)
values = Reference(ws, min_col=5, max_col=5, min_row=2, max_row=ws.max_row)
chart.add_data(values, titles_from_data=False)
chart.set_categories(categories)

ws.add_chart(chart, "A10")
wb.save("charts.xlsx")

# from openpyxl.chart import BarChart, Series


# Create combined Column Chart
chart2 = BarChart()
chart2.title = "Price vs Units Sold"
chart2.x_axis.title = "Product"
chart2.y_axis.title = "Value"

price_data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
units_data = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row)
categories = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)

chart2.add_data(price_data, titles_from_data=True)
chart2.add_data(units_data, titles_from_data=True)
chart2.set_categories(categories)

ws.add_chart(chart2, "J1")
wb.save("charts.xlsx")


# Exporting pandas charts ----------------------

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage


df = pd.read_excel('product_data.xlsx')

df[['product_price', 'Units Sold']].plot(kind='bar', figsize=(4, 3))
plt.title("Price and Units Sold per Product")
plt.ylabel("Value")
plt.xticks(rotation=15)
plt.tight_layout()

plt.savefig("chart.png", bbox_inches="tight")

ws2 = wb.create_sheet(title="pandas chart")

img = XLImage("chart.png")
ws2.add_image(img, "B2")

wb.save("charts.xlsx")