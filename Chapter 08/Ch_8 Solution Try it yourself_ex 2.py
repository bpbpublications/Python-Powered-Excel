# Exercise 2a -----------------------------------------------

from openpyxl import Workbook

# Create a workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Maintenance Log"

# Define headers and maintenance data
headers = ["Date", "Machine ID", "Issue", "Action Taken", "Technician", "Downtime (hrs)"]

entries = [
    ["7/1/2025", "MX100", "Vibration above threshold", "Bearing replaced", "Anil", 2.5],
    ["7/2/2025", "MX101", "Overheating", "Coolant replaced", "Ravi", 1.8],
    ["7/3/2025", "MX100", "Oil leak", "Seal fixed", "Anil", 1.2],
    ["7/4/2025", "MX102", "Sensor failure", "Sensor replaced", "Priya", 2],
    ["7/5/2025", "MX100", "Vibration above threshold", "Realigned rotor", "Karan", 2.7],
    ["7/6/2025", "MX100", "Vibration above threshold", "Bearing replaced", "Anil", 2.5],
    ["7/7/2025", "MX101", "Overheating", "Coolant replaced", "Ravi", 1.8],
    ["7/8/2025", "MX100", "Oil leak", "Seal fixed", "Anil", 1.2],
    ["7/9/2025", "MX102", "Sensor failure", "Sensor replaced", "Priya", 2],
    ["7/10/2025", "MX100", "Vibration above threshold", "Realigned rotor", "Karan", 2.7],
    ["7/11/2025", "MX101", "Oil leak", "Seal replaced", "Ravi", 1.5],
    ["7/12/2025", "MX103", "Electrical fault", "Rewired panel", "Priya", 2.2],
    ["7/13/2025", "MX100", "Overheating", "Cleaned radiator", "Karan", 2],
    ["7/14/2025", "MX101", "Sensor failure", "Sensor recalibrated", "Anil", 1.6],
    ["7/15/2025", "MX102", "Vibration above threshold", "Balanced rotor", "Priya", 2.1],
    ["7/16/2025", "MX103", "Belt slippage", "Belt tightened", "Ravi", 1.3],
    ["7/17/2025", "MX101", "Oil leak", "Seal replaced", "Karan", 1.4],
    ["7/18/2025", "MX102", "Temperature spike", "Coolant flushed", "Anil", 2.3],
    ["7/19/2025", "MX100", "Motor failure", "Motor replaced", "Priya", 3],
    ["7/20/2025", "MX103", "Vibration above threshold", "Bearing replaced", "Karan", 2.6],
]


# Populate worksheet
ws.append(headers)
for row in entries:
    ws.append(row)

wb.save('Maintenance_Log.xlsx') 

# Exercise 2b ---------------------------------------------
# Restart session 

from openpyxl import load_workbook

file = "Maintenance_Log.xlsx"
wb = load_workbook(file)
ws = wb.active

row_count = 0

for row in ws.values:
    if row_count == 0:
        print(f"Headers: {row}")
    elif row_count >= 0:
        print(f"Row {row_count}: {row}")
    row_count += 1

print(f"\nTotal rows: {row_count}")

# 2c -----------------------
ws.row_dimensions[1].height = 25


# 2d -----------------------
from openpyxl.styles import Font

header_font = Font(name="Agency FB", 
                   size=14, bold=True, italic=True)

for cell in ws[1]:
    cell.font = header_font
    
# 2e -------------------------
from openpyxl.styles import Alignment

ws.insert_rows(1, amount=2)

ws["A1"] = "Sample Data"
ws.merge_cells("A1:F1")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

# 2f ------------------------
from openpyxl.styles import PatternFill

# threshold > 2 hours
threshold = 2

fill_red = PatternFill(start_color="FFC7CE", 
                       end_color="FFC7CE", fill_type="solid")

for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    downtime = row[5].value
    if downtime > threshold:
        for cell in row:
            cell.fill = fill_red

# 2 g ----------------------------

from openpyxl.styles import Border, Side

tech_count = {}

for row in ws.iter_rows(min_row=4, values_only=True):
    tech = row[4]
    if tech in tech_count:
        tech_count[tech] += 1
    else:
        tech_count[tech] = 1

max_tech = max(tech_count, key=tech_count.get)

border = Border(
    left=Side(style="thick"),
    right=Side(style="thick"),
    top=Side(style="thick"),
    bottom=Side(style="thick")
)

for row in ws.iter_rows(min_row=4):
    if row[4].value == max_tech:
        for cell in row:
            cell.border = border
            
            
# 2h -----------------------------

report_ws = wb.create_sheet("Report")
report_ws.sheet_properties.tabColor = "1072BA"

# 2 i -----------------------------

# There are various ways to compute these KPIs
# You can use pandas too

downtime_per_machine = {}
downtime_per_issue = {}
downtime_per_tech = {}
count_per_issue = {}
count_per_tech = {}


for row in ws.iter_rows(min_row=4, values_only=True):
    machine = row[1]
    issue = row[2]
    tech = row[4]
    downtime = row[5]

    # Total downtime per machine
    if machine in downtime_per_machine:
        downtime_per_machine[machine] += downtime
    else:
        downtime_per_machine[machine] = downtime

    # Avg downtime per issue
    if issue in downtime_per_issue:
        downtime_per_issue[issue] += downtime
        count_per_issue[issue] += 1
    else:
        downtime_per_issue[issue] = downtime
        count_per_issue[issue] = 1

    # Avg downtime per technician
    if tech in downtime_per_tech:
        downtime_per_tech[tech] += downtime
        count_per_tech[tech] += 1
    else:
        downtime_per_tech[tech] = downtime
        count_per_tech[tech] = 1


# Compute averages
avg_downtime_per_issue = {}
for issue in downtime_per_issue:
    avg_downtime_per_issue[issue] = downtime_per_issue[issue] / count_per_issue[issue]

avg_downtime_per_tech = {}
for tech in downtime_per_tech:
    avg_downtime_per_tech[tech] = downtime_per_tech[tech] / count_per_tech[tech]

# Most frequent issue
most_freq_issue = None
max_issue_count = 0

for issue in count_per_issue:
    if count_per_issue[issue] > max_issue_count:
        max_issue_count = count_per_issue[issue]
        most_freq_issue = issue

# 2 j -----------------------------

# # Create sheet (done already)
# report_ws = wb.create_sheet("Reports")
# report_ws.sheet_properties.tabColor = "1072BA"

# Title
report_ws["A1"] = "Maintenance KPI Report"
report_ws["A1"].font = Font(size=16, bold=True)

# Machine Downtime

report_ws["A3"] = "Total Downtime per Machine"
report_ws["A3"].font = Font(bold=True)

report_ws["A4"] = "Machine"
report_ws["B4"] = "Total Downtime"

row = 5
for machine in downtime_per_machine:
    report_ws.cell(row=row, column=1, value=machine)
    report_ws.cell(row=row, column=2, value=downtime_per_machine[machine])
    row += 1

# Issue Metrics

report_ws["D3"] = "Issue-wise Metrics"
report_ws["D3"].font = Font(bold=True)

report_ws["D4"] = "Issue"
report_ws["E4"] = "Total Downtime"
report_ws["F4"] = "Count"
report_ws["G4"] = "Avg Downtime"

row = 5
for issue in downtime_per_issue:
    report_ws.cell(row=row, column=4, value=issue)
    report_ws.cell(row=row, column=5, value=downtime_per_issue[issue])
    report_ws.cell(row=row, column=6, value=count_per_issue[issue])
    report_ws.cell(row=row, column=7, value=round(avg_downtime_per_issue[issue], 1))
    row += 1

#  Technician Metrics

report_ws["A14"] = "Technician-wise Metrics"
report_ws["A14"].font = Font(bold=True)

report_ws["A15"] = "Technician"
report_ws["B15"] = "Total Downtime"
report_ws["C15"] = "Count"
report_ws["D15"] = "Avg Downtime"

row = 16
for tech in downtime_per_tech:
    report_ws.cell(row=row, column=1, value=tech)
    report_ws.cell(row=row, column=2, value=downtime_per_tech[tech])
    report_ws.cell(row=row, column=3, value=count_per_tech[tech])
    report_ws.cell(row=row, column=4, value=avg_downtime_per_tech[tech])
    row += 1

# Key Insight

report_ws["F14"] = "Key Insight"
report_ws["F14"].font = Font(bold=True)

report_ws["F15"] = "Most Frequent Issue"
report_ws["G15"] = most_freq_issue

report_ws["F16"] = "Issue Count"
report_ws["G16"] = max_issue_count

# 2 k ------------------------------

from openpyxl.chart import BarChart, Reference

# Add bar chart to analyze downtime
chart = BarChart()
chart.title = "Machine Downtime (hrs)"
chart.x_axis.title = "Machine ID"
chart.y_axis.title = "Downtime (hrs)"
data = Reference(report_ws, min_col=2, min_row=5, max_row=8)
labels = Reference(report_ws, min_col=1, min_row=5, max_row=8)
chart.add_data(data, titles_from_data=False)
chart.set_categories(labels)

chart.y_axis.majorGridlines = None
chart.x_axis.majorGridlines = None

report_ws.add_chart(chart, "K3")

wb.save('Business_KPIs.xlsx')

# 2 l ------------------------------
wb.template = True
wb.save("template_file.xltx")