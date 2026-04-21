from openpyxl import Workbook 

# Creating workbook -------------------------
wb = Workbook() 
ws = wb.active
ws.title = 'products'

# Header list
headers = ['product_id', 'product_name', 'product_category', 'product_price']

ws.append(headers)
# wb.save('product_data.xlsx')


# # Populate headers dynamically -------------------------

# for col_num, header in enumerate(headers, start=1):
#     cell = ws.cell(row=1, column=col_num)
#     cell.value = header
    
# wb.save('product_data.xlsx')

# Adding Data from a List of Lists -------------------------

data = [
    [101, 'Wireless Mouse', 'Electronics', 25.99],
    [102, 'Yoga Mat', 'Fitness', 15.50],
    [103, 'Notebook', 'Stationery', 3.99]
]

# Start writing from row 2 (after header)
for row in data:
    ws.append(row)


# Preview data by printing each row -----------
# ws_data = ws.values

# for row in ws_data:
#     print(row) 

# Creating a function to print each row 

def print_rows():
     for row in ws.iter_rows(values_only=True):
         print(row)

print_rows()

# Writing data using cell level control  -------------------------
data = [
    [104, 'Pen', 'Stationery', 10],
    [105, 'USB Cable', 'Electronics', 150]
]

for r_idx, row in enumerate(data, start=ws.max_row + 1):
    for c_idx, val in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=val)
        
print_rows()
        
# Adding data from a dictionary -------------------------

products = [
    {'id': 106, 'name': 'Chair', 'category': 'Furniture', 'price': 45.50},
    {'id': 107, 'name': 'Water Bottle', 'category': 'Fitness', 'price': 10.00}
]

for product in products:
    ws.append([
        product['id'],
        product['name'],
        product['category'],
        product['price']
    ])

print_rows()

wb.save('product_data.xlsx')
        
# Writing data to column -------------------------        

ws['E1'] = 'Units Sold'
quantity = [5, 8, 15, 30, 6, 10, 20]

for i, status in enumerate(quantity , start=2):
    ws[f'E{i}'] = status

wb.save('product_data.xlsx')

# availability = ["In Stock", "Out of Stock", "In Stock", "In Stock", "Out of Stock"]
# ws['F1'] = 'Availability'

# for i, status in enumerate(availability , start=2):
#     ws[f'F{i}'] = status

# Add a derived column, calculated from other columns -------------------------
ws['F1'] = 'Revenue'
for i in range(2, ws.max_row + 1):
    price = ws[f'D{i}'].value
    units_sold = ws[f'E{i}'].value
    ws[f'F{i}'] = price*units_sold

wb.save('product_data_derived_column.xlsx')