from openpyxl import load_workbook
file = "product_data.xlsx"
wb = load_workbook(file)
ws = wb.active
 

# Freeze all rows above row 2 (i.e. top row) and columns to the left of column B(i.e. first columns)

ws.freeze_panes = 'B2'

# # Freeze just the top row (row 1)
# ws.freeze_panes = 'A2'
# # Freeze just the first column (column A)
# ws.freeze_panes = 'B1'

# Remove freeze panes
ws.freeze_panes = None


ws.auto_filter.ref = ws.dimensions # AutoFilter to the entire used range of the worksheet
ws.auto_filter.ref = None # Remove the AutoFilter from the worksheet


ws.insert_rows(1, 2)
ws['A1'] = 'Sample Data'

ws.row_dimensions[1].height = 20

ws.dimensions # 'A1:F10'


# Alignment Menu of Excel [merge/wrap, cell alignment] -----------------

ws.merge_cells('A1:F1')
# ws.unmerge_cells('A1:F1')

from openpyxl.styles import Alignment
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')


ws['H3'] = 'OpenPyXl is an awesome library that helps me automate Excel workflow.'
ws['H3'].alignment = Alignment(wrap_text=True)
ws.column_dimensions['H'].width = 15

wb.save("data_formatting.xlsx")

# Font menu of Excel [Styles-, font, size, color, fill, bold/italic etc] -------------

from openpyxl.styles import Font

ws['A1'].font = Font(size=16, bold=True, color='FF1F497D')
ws['H3'].font = Font(italic=True, name = 'Agency FB')

wb.save("data_formatting.xlsx")

# Using NamedStyle ----------------------------------------

from openpyxl.styles import NamedStyle, PatternFill, Border, Side

# Define a named style
custom_style = NamedStyle(name="custom_style")

custom_style.font = Font(size=12, bold=True, color="FF0000")
custom_style.fill = PatternFill(start_color="ADD8E6", fill_type="solid")
custom_style.border = Border(
    left=Side(style='medium'),
    right=Side(style='mediumDashed'),
    top=Side(style='double'),
    bottom=Side(style='slantDashDot')
)

# Register style to workbook
wb.add_named_style(custom_style)

# Apply style to another cell

ws['A3'].style = "custom_style"

# Save again
wb.save("formatting_NamedStyle.xlsx")

# To apply to 4th row
for cell in ws[4]:
    cell.style = "custom_style"