from openpyxl import Workbook

wb = Workbook() # Creating a new Excel Workbook
ws = wb.active # Accessing active sheet (the first sheet)

ws.title = 'products' # Giving this sheet a title

# Assigning some values to cells
ws['A1'] = 'product_id' # Alternately, you can use .cell method as follows
# ws.cell(row=1, column=1, value = 'product_id')

ws['B1'] = 'product_name'
ws['C1'] = 'product_category'
ws['D1'] = 'product_price'

wb.save('business_data.xlsx') # saving as xlsx file

ws.max_row

# To append data (new row)
new_row = ('P001', 'GrandI10_Sportz','Car',7.8)
ws.append(new_row)
    
wb.save('business_data.xlsx')

ws.delete_rows(ws.max_row, 1) # row number, number of rows to delete

wb.save('business_data.xlsx')

# Create a second sheet by title
wb.create_sheet(title="customers")

# Accessing the sheet by it's name
ws2 = wb['customers']
# or using index as below
# ws2 = wb.worksheets[1]

ws2['A1'] = 'customer_id'
ws2['B1'] = 'region' 

# Creating another sheet
wb.create_sheet('delivery')

# Accessing the sheet by it's index (index starts from 0)
ws3 = wb.worksheets[2]
ws3

# Value assignment to cells at location (row, col)
ws3.cell(1, 1, 'order_id')
ws3.cell(1, 2, 'order_date') 
ws3.cell(1, 3, 'proposed_delivery_date')
ws3.cell(1, 4, 'actual_delivery_date')

# Creating a sheet at index 3
ws4 = wb.create_sheet('Analytics_Report', 4)
ws4.sheet_properties.tabColor = '1072BA'
# Colors must be a RGB hex values

# Again trying to create sheet with same name, name will get suffix 1
ws5 = wb.create_sheet('Analytics_Report')

# Deleting sheet named 'Analytics_Report1'
wb.remove(wb['Analytics_Report1'])

# Creating copy of the sheet 'products'

products_copy = wb.copy_worksheet(wb['products'])
products_copy.title = 'products_copy'

wb.worksheets
wb.remove(wb['products_copy'])


# insert a sheet named 'marketing' at index 3, i.e., 4th sheet
wb.create_sheet('marketing', 3)

# Moving 'delivery' sheet to one position earlier
wb.move_sheet(wb['delivery'], offset=-1)
wb.worksheets # Sheet order changed

# Move the sheet 'products' to the last position
wb.move_sheet(wb['products'], offset=len(wb.worksheets) - 1)

wb.move_sheet(wb['Analytics_Report'], len(wb.worksheets) - 1)

# list of all sheets name
wb.worksheets

# inserting 2 rows before the 1st row in products sheet
ws
ws.insert_rows(idx = 0, amount=2)

# inserting 1 col before the 3rd column 
ws.insert_cols(idx=2)

ws['A3'] = 'ID' # renaming col

wb.save('business_data.xlsx')