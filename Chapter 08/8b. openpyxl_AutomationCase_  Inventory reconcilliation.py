from openpyxl import load_workbook

wb = load_workbook("Inventory_Comparison.xlsx")
ws1 = wb['ERP_Inventory']
ws2 = wb['Warehouse_Count']

for row in range(2, ws1.max_row + 1):
    item_id = ws1[f"A{row}"].value
    erp = ws1[f"D{row}"].value
    warehouse = ws2[f"D{row}"].value
    if erp != warehouse:
        print(f"Item ID {item_id}:\n System-recorded stock levels={erp}, Physically counted stock at warehouse={warehouse}\n*****")