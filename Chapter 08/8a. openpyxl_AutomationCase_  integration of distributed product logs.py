# Uses-Case: Automated integration of distributed product logs

# This script: 
# •	Loads all departmental files (.xlsx files) from a folder
# •	Extracts and appends their content into a master workbook, e.g., consolidated_log.xlsx


from openpyxl import load_workbook, Workbook
import os

folder = "department_logs"
master_wb = Workbook()
master_ws = master_wb.active
master_ws.title = "All Departments"
master_ws.append(['product_id', 'product_name', 'product_category', 'product_price', 'Units Sold'])

for file in os.listdir(folder):
    if file.endswith('.xlsx'):
        wb = load_workbook(os.path.join(folder, file))
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            master_ws.append(row)

master_wb.save("consolidated_log.xlsx")

# Alternate way_Using list comprehension 
files = [f for f in os.listdir(folder) if f.endswith(".xlsx")]
for file in files:
    wb = load_workbook(os.path.join(folder, file))
    ws = wb.active
    print(f"{file}: {ws.max_row} rows")
    for row in ws.iter_rows(min_row=2, values_only=True):
        master_ws.append(row)

master_wb.save("consolidated_log.xlsx")