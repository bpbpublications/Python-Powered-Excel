# Data Validation --------------------------------
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


# RULE 1: price >0

wb = load_workbook('product_data.xlsx')
ws = wb.active

# Create validation rule: Only allow decimal values > 0
dv_price = DataValidation(type="decimal", 
                          operator="greaterThan", 
                          formula1="0", showErrorMessage=True)

dv_price.error = "Price must be a positive number."
dv_price.errorTitle = "Invalid Price"

# Apply to product_price column (D2:D7)
# ws.add_data_validation(dv_price)
# dv_price.add("D2:D7")

# Add the validation to the column dynamically
# last_row = ws.max_row
# range_str = f"D2:D{last_row}"
# ws.add_data_validation(dv_price)
# dv_price.add(range_str)

max_excel_rows = 1048576  # Excel's row limit
dv_price.add(f"D2:D{max_excel_rows}")

ws.add_data_validation(dv_price)
wb.save('products_data_validated.xlsx')

ws['D12'] = -87

# Rule 2: Category from a given list -----------

# Create data validation for a dropdown list
dv_cat = DataValidation(type="list", 
                        formula1='"Electronics,Fitness,Stationery"', 
                        allow_blank=False)
dv_cat.prompt = "Choose a value from the list"
dv_cat.error = "Invalid entry. Please choose from the list."

# Apply the data validation to a range of cells
ws.add_data_validation(dv_cat)
dv_cat.add(f"C2:C{max_excel_rows}")

wb.save('products_data_validated1.xlsx')

# Rule 3:  Text Length Validation (e.g., < 0 characters)
    
# Create text length validation rule
dv_text = DataValidation(
    type="textLength",
    operator="lessThan",
    formula1="10",
    showErrorMessage=True
)
dv_text.error = "Too long! Text must be under 10 characters."
dv_text.errorTitle = "Text Length Error"

# Apply validation 
dv_text.add(f"B2:B{max_excel_rows}")
ws.add_data_validation(dv_text)

wb.save("text_length_validation.xlsx")

    
# Conditional Formatting ------------------------------

# Highlight cells in “Units Sold” > 10
    
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

# Define formatting rule: highlight with green fill
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
rule = CellIsRule(operator='greaterThan', formula=['10'], fill=green_fill)

# Apply to "Units Sold" column (E column)
# ws.conditional_formatting.add("E2:E8", rule)
last_row = ws.max_row
E_range_str = f"E2:E{last_row}"
ws.conditional_formatting.add(E_range_str, rule)

wb.save('conditional_formatting.xlsx')