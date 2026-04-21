from openpyxl import load_workbook
file = "product_data.xlsx"
wb = load_workbook(file)
ws = wb.active

# Preview data by printing each row -----------

def print_rows():
     for row in ws.iter_rows(values_only=True):
         print(row)

print_rows()
# # ------------------------------------------------------

# Excel formula for a derived column 'Revenue'
ws['F1'] = 'Revenue'

for i in range(2, ws.max_row + 1):
    price = f'D{i}'
    units_sold = f'E{i}'
    ws[f'F{i}'] = f'={price}*{units_sold}'
    
wb.save('writing_excel_formula.xlsx')    

# Working with formula

ws['H1'] = 'Total Revenue'
ws['H2'] = '=SUM(F2:F8)'

# ws['I1'] = 'Average Revenue'
# ws['I2'] = '=AVERAGE(F2:F7)'

wb.save('writing_excel_formula.xlsx')   

# Or doing it dynamically
last_row = ws.max_row

ws[f'E{last_row + 1}'] = 'Total Revenue'
ws[f'F{last_row + 1}'] = f'=SUM(F2:F{last_row})'
sum_formula = f'=SUM(F2:F{last_row})'

# alternate way
# ws.cell(row=last_row + 1, column=5, value='Total Revenue')
# ws.cell(row=last_row + 1, column=6, value=sum_formula)

wb.save('writing_excel_formula.xlsx')  

# AVERAGE formula
mean_formula = f'=AVERAGE(F2:F{last_row})'
ws.cell(row=last_row + 2, column=5, value='Mean Revenue')
ws.cell(row=last_row + 2, column=6, value=mean_formula)
wb.save('writing_excel_formula.xlsx')  

# Other available formulae
from openpyxl.utils import FORMULAE
FORMULAE

type(FORMULAE)
len(FORMULAE)

# To check if 'TEXTJOIN' formula is supported
'TEXTJOIN' in FORMULAE # False

ws['I1'] = 'TEXTJOIN() Function'
ws['I2'] = '=TEXTJOIN(": ", TRUE, B2:C2)'  # Might show #NAME error in Excel

ws['J1'] = 'using _xlfn prefix'
ws['J2'] = '=_xlfn.TEXTJOIN(": ", TRUE, B2:C2)'  # Now works
wb.save('writing_excel_formula.xlsx')

# Hiding some columns to view ot better
for col_letter in ['A', 'D', 'E', 'F', 'H']:
    ws.column_dimensions[col_letter].hidden = True
    
# Array Formulas --------------------

from openpyxl.worksheet.formula import ArrayFormula

ws['I4']  = "Transposed Data"
ws['I6'].value = ArrayFormula('I6:J7', '=TRANSPOSE(I1:J2)')

# Hiding some columns to view it better
for col_letter in ['B', 'C']:
    ws.column_dimensions[col_letter].hidden = True
    
wb.save('writing_excel_formula.xlsx')

# if you want to clean-up 
for row in ws.iter_rows(min_row=1, max_row=7, min_col=8, max_col=10):
    for cell in row:
        print(cell)
        cell.value = None
        
# Unhiding some columns to view it better
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'H']:
    ws.column_dimensions[col_letter].hidden = False

# To hide a range of columns from 'A' to 'H' using grouping
ws.column_dimensions.group(start='A', end='H', hidden=False)
wb.save('grouped_hidden_sales.xlsx')