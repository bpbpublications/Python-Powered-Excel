import os
from openpyxl import load_workbook

# Folder containing all budget spreadsheets
folder_path = 'department_budgets'

# Column and sheet to scan
target_column = 'C'  # Adjust based on your layout
sheet_name = 'summary'  # Or use wb.active if sheet names vary

# Scan all files
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        for row in range(2, ws.max_row + 1):  # Skip header
            cell = ws[f'{target_column}{row}']
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                print(f"🔴 Deficit in {filename} at {cell.coordinate}: {cell.value}")
                break  # Stop after first deficit found in this file