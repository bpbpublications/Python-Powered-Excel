# Loading data ---------------------
from openpyxl import load_workbook
file = "mydata.xlsx"
wb = load_workbook(file)

# Accessing Sheets ---------------------

wb.sheetnames # All sheets names

ws = wb.active # or active sheet/last opened sheet
print(ws.title)

ws = wb['iris'] # To fetch sheet by sheet name
print(ws)
# OR To load sheet by using sheet number {index starts at 0}
ws = wb.worksheets[0]
print(ws)

# Sheet dimensions---------------------

print(ws.max_row), print(ws.max_column)

# Accessing Data ---------------------

# A single cell
ws['A1']
ws.cell(1, 1)

# a range of cells
ws['A1':'E1'] # 2-d tuple
ws['A1':'E1'][0] # 1-d tuple

# Single Column     
colC = ws['C']
colC


# Column Range
col_range = ws['B:D'] # All the rows for columns B up till D
col_range


# Single Row  
row5 = ws[5] 
row5

# Row Range
row_range = ws[5:10] # All the columns for rows 5 to 10
row_range

# Reading Data ---------------------

# Accessing Data from a single cell
ws['A1'].value
type(ws['A1'].value)

ws.cell(1, 1).value

# From ranges_Using loops

for cell in ws['A1':'E1'][0]:
    print(cell.value)
    
for row in ws['B2:D4']:  # 2D tuple of cells
    print(row)
    for cell in row:
        print(cell.coordinate, "=", cell.value)
        
# From a column    
colC = ws['C']

for cell in ws['C']:  # Column C
    print(cell.coordinate, "=", cell.value)
   
# list comprehension
[(cell.coordinate, cell.value) for cell in ws['C']]    
 
# From a row    
row5 = ws[5]

for cell in ws[5]:  # Row 5
    print(cell.coordinate, cell.value)
 
# list comprehension
[(cell.coordinate, cell.value) for cell in ws[5]]      
    
# Alternately
[ws.cell(row=i,column=3).value for i in range(2,12)] # Col C, selected rows
[ws.cell(row=5,column=i).value for i in range(1,ws.max_column+1)]
   
 
# From range of columns
    
col_range = ws['D:E']; col_range # 2 d tuple
col_range[0] # gives D column
col_range[0][0:5] # First 5 entries of d column

for cell in ws['D:E'][0][0:5]:
    print(cell.coordinate, cell.value)

row_range = ws[5:10]

# Better Approach
for col_cells in ws.iter_cols(min_col=2, max_col=4):  # B to D
    for cell in col_cells:
        print(cell.coordinate, cell.value)

# Ranges such as values from columns B to D for rows 2 to 5:


# Loop through each row and print values (row-wise)
    
for row in ws.iter_rows(min_row=2, max_row=5, 
                            min_col=2, max_col=4, 
                            values_only=True):
    print(row)

# Loop through each column and print values (column-wise)

for col in ws.iter_cols(min_row=2, max_row=5, 
                            min_col=2, max_col=4, 
                            values_only=True):
    print(col)
    
    
# Dynamic ranges (all rows, starting from 2) ---------------------
   
max_row = ws.max_row; max_row
max_column = ws.max_column; max_column


for row in ws.iter_rows(min_row=2, max_row=max_row, 
                            min_col=2, max_col=4, 
                            values_only=True):
    print(row)

# Skip empty cells, if any
for row in ws.iter_rows(min_row=2, max_row=5, 
                        min_col=2, max_col=4):
    for cell in row:
        if cell.value is not None:  # Skip empty cells
            print(cell.value)    
    
# Non-Consecutive Columns ---------------------
cols = ['A', 'C', 'E']

for col in cols:
    for cell in ws[col]:
        print(cell.coordinate, cell.value)

for col in cols:
    for cell in ws[col][:5]:
        print(cell.coordinate, cell.value)
    
ws.max_row
ws.max_column
    
# To read entire sheet datab------------------------

# Get the generator object containing worksheet values 
ws_data = ws.values

# Preview data by printing each row (optional debugging step)
for row in ws_data:
    print(row)  # Each row is a tuple

ws_data = ws.values
sheet_data = [row for row in ws_data] # list

# Reset the generator to read again from the beginning
ws_data = ws.values

# Extract the header row (column names)
next(ws_data)  # Take all columns in the header row

# Create a new generator object to get data excluding headers
ws_data = ws.values  # Reset again to fetch from start
next(ws_data)        # Skip header row

# Write to pandas DataFrame ---------------------
ws_data = ws.values  # Reset again to fetch from start

import pandas as pd
# Convert the worksheet data into a DataFrame using the extracted headers
data = pd.DataFrame(ws_data, columns = next(ws_data))

wb.close()

# read_only() mode ----------------------------

# The read_only=True argument in openpyxl.load_workbook() tells 
# the library to load the workbook in read-only mode, 
# which is super handy if you're working with large Excel files 
# or you don't need to modify the data.

from openpyxl import load_workbook

wb = load_workbook("mydata.xlsx", read_only=True)
ws = wb.active

ws['A1'].value = "Modified value" # AttributeError: Cell is read only

try:
    ws['A1'].value = "Modified value"
except AttributeError as e:
    print("You cannot modify cells in read-only mode.")
    print("Error:", e)

wb.close()
# NOTE: The workbook opened with read_only mode must be 
# explicitly closed with the close() method

# Showing formula Vs showing values while loading data -------------

file = "data_only_option.xlsx"   

# shows values (not formula)
wb = load_workbook(file, data_only=True) # shows values (not formula)
ws = wb.active 


for row in ws.iter_rows(values_only=True):
            print(row)

            
# Shows actual formula as string 
wb = load_workbook(file, data_only=False) # Shows actual formula as string
ws = wb.active 

for row in ws.iter_rows(values_only=True):
            print(row)