# For demo purposes, I've written simple functions
# You need to improve these functions to handle edge cases


from openpyxl import load_workbook, Workbook
from datetime import datetime, date

def clean_text(value):
    if isinstance(value, str):
        return value.strip().lower()  # Remove spaces, normalize case
    return value

def clean_price(value):
    try:
        return float(value)
    except:
        return None  # Mark invalid numbers

def clean_date(value):
    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()
    elif isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y.%m.%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except:
                continue
    return None  # Unknown format

# Load raw workbook
wb = load_workbook("Inconsistent_Upload.xlsx")
ws = wb["Raw_Upload"]

# Create cleaned workbook
cleaned_wb = Workbook()
cleaned_ws = cleaned_wb.active
cleaned_ws.title = "Cleaned_Data"

# Copy headers
headers = [cell.value for cell in ws[1]]
cleaned_ws.append(headers)

# Clean each row
for row in ws.iter_rows(min_row=2, values_only=True):
    product_id = row[0] if row[0] else "UNKNOWN"
    product_name = clean_text(row[1])
    category = clean_text(row[2])
    price = clean_price(row[3])
    date_added = clean_date(row[4])
    
    cleaned_ws.append([product_id, product_name, category, price, date_added])

# Save cleaned file
cleaned_wb.save("Cleaned_Data.xlsx")
