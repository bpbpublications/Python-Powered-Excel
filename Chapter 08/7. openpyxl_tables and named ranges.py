from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = load_workbook('product_data.xlsx')
ws = wb.active

# Get the last row
last_row = ws.max_row
# Get the last column letter
last_col_letter = get_column_letter(ws.max_column)

# Creating Excel Table --------------------------
# Define the table range (adjust for your actual data)
table_ref = f"A1:{last_col_letter}{last_row}"
table = Table(displayName="ProductsTable", ref=table_ref)

# Optional: add style to table
style = TableStyleInfo(name="TableStyleMedium9", 
                       showRowStripes=True)
table.tableStyleInfo = style

# Add table to worksheet
ws.add_table(table)

wb.save('products_with_table.xlsx')

ws.tables # name of tables
ws.tables.keys()
ws.tables.items() #  list of table name and their ranges

# https://openpyxl.readthedocs.io/en/stable/worksheet_tables.html

# Named Range ----------------------------------

from openpyxl.workbook.defined_name import DefinedName

# Define a named range for product prices (D2:D8)
price_range = DefinedName('ProductPrices', 
                          attr_text='Sheet1!$D$2:$D$8')
wb.defined_names.add(price_range)

wb.defined_names

wb.save('products_with_named_range.xlsx')


# Tables auto expand -------------------------------
# Open products_with_named_range.xlsx in Excel and
# add a new row to table before running the following code:
    
wb = load_workbook('products_with_named_range.xlsx')

ws = wb.active

ws.tables.items() #  list of table name and their ranges
# ('ProductsTable', 'A1:F9')] It shows 9th row is added to table